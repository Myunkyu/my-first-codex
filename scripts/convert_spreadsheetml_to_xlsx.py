from __future__ import annotations

from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SS = "urn:schemas-microsoft-com:office:spreadsheet"
NS = {"ss": SS}

SOURCE = Path("artifacts/quantum_reference_database_20260704.xml")
OUTPUT = Path("artifacts/양자분야_시장기술정책표준화_참고자료_DB_20260704.xlsx")


def attr(element: ET.Element, name: str, default: str | None = None) -> str | None:
    return element.get(f"{{{SS}}}{name}", default)


def parse_color(value: str | None, default: str | None = None) -> str | None:
    if not value:
        return default
    return value.replace("#", "").upper()


def build_style_map(root: ET.Element) -> dict[str, dict]:
    styles: dict[str, dict] = {}
    styles_node = root.find("ss:Styles", NS)
    if styles_node is None:
        return styles

    for style in styles_node.findall("ss:Style", NS):
        style_id = attr(style, "ID")
        if not style_id:
            continue
        cfg: dict = {}

        font_node = style.find("ss:Font", NS)
        if font_node is not None:
            cfg["font"] = Font(
                name=attr(font_node, "FontName") or "Malgun Gothic",
                size=float(attr(font_node, "Size") or 10),
                bold=attr(font_node, "Bold") == "1",
                italic=attr(font_node, "Italic") == "1",
                underline="single" if attr(font_node, "Underline") else None,
                color=parse_color(attr(font_node, "Color")),
            )

        fill_node = style.find("ss:Interior", NS)
        if fill_node is not None and attr(fill_node, "Pattern") == "Solid":
            color = parse_color(attr(fill_node, "Color"), "FFFFFF")
            cfg["fill"] = PatternFill(fill_type="solid", fgColor=color)

        align_node = style.find("ss:Alignment", NS)
        if align_node is not None:
            h_map = {"Center": "center", "Left": "left", "Right": "right"}
            v_map = {"Center": "center", "Top": "top", "Bottom": "bottom"}
            cfg["alignment"] = Alignment(
                horizontal=h_map.get(attr(align_node, "Horizontal")),
                vertical=v_map.get(attr(align_node, "Vertical"), "top"),
                wrap_text=attr(align_node, "WrapText") == "1",
            )

        borders_node = style.find("ss:Borders", NS)
        if borders_node is not None:
            sides = {"Left": Side(), "Right": Side(), "Top": Side(), "Bottom": Side()}
            for b in borders_node.findall("ss:Border", NS):
                pos = attr(b, "Position")
                if pos in sides:
                    sides[pos] = Side(
                        style="thin" if attr(b, "LineStyle") == "Continuous" else None,
                        color=parse_color(attr(b, "Color"), "D9E1F2"),
                    )
            cfg["border"] = Border(
                left=sides["Left"], right=sides["Right"],
                top=sides["Top"], bottom=sides["Bottom"],
            )

        styles[style_id] = cfg
    return styles


def apply_style(cell, cfg: dict) -> None:
    if "font" in cfg:
        cell.font = cfg["font"]
    if "fill" in cfg:
        cell.fill = cfg["fill"]
    if "alignment" in cfg:
        cell.alignment = cfg["alignment"]
    if "border" in cfg:
        cell.border = cfg["border"]


def main() -> None:
    tree = ET.parse(SOURCE)
    root = tree.getroot()
    style_map = build_style_map(root)

    wb = Workbook()
    wb.remove(wb.active)

    for worksheet in root.findall("ss:Worksheet", NS):
        sheet_name = attr(worksheet, "Name") or "Sheet"
        ws = wb.create_sheet(sheet_name[:31])
        ws.sheet_view.showGridLines = False

        table = worksheet.find("ss:Table", NS)
        if table is None:
            continue

        for idx, col in enumerate(table.findall("ss:Column", NS), start=1):
            width = float(attr(col, "Width") or 64)
            ws.column_dimensions[get_column_letter(idx)].width = max(8, width / 7.2)

        row_number = 0
        for row_node in table.findall("ss:Row", NS):
            row_number += 1
            height = attr(row_node, "Height")
            if height:
                ws.row_dimensions[row_number].height = float(height)

            col_number = 0
            for cell_node in row_node.findall("ss:Cell", NS):
                explicit_index = attr(cell_node, "Index")
                if explicit_index:
                    col_number = int(explicit_index) - 1
                col_number += 1

                cell = ws.cell(row=row_number, column=col_number)
                data_node = cell_node.find("ss:Data", NS)
                if data_node is not None:
                    data_type = attr(data_node, "Type") or "String"
                    text = data_node.text or ""
                    if data_type == "Number":
                        try:
                            cell.value = float(text)
                        except ValueError:
                            cell.value = text
                    else:
                        cell.value = text

                style_id = attr(cell_node, "StyleID")
                if style_id and style_id in style_map:
                    apply_style(cell, style_map[style_id])

                href = attr(cell_node, "HRef")
                if href:
                    cell.hyperlink = href
                    if not cell.value:
                        cell.value = "원문 열기"

                merge_across = int(attr(cell_node, "MergeAcross") or 0)
                merge_down = int(attr(cell_node, "MergeDown") or 0)
                if merge_across or merge_down:
                    ws.merge_cells(
                        start_row=row_number,
                        start_column=col_number,
                        end_row=row_number + merge_down,
                        end_column=col_number + merge_across,
                    )
                    col_number += merge_across

        if sheet_name == "01_참고자료_DB":
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:N{ws.max_row}"
        elif sheet_name == "04_지속_모니터링":
            ws.freeze_panes = "A3"

        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    wb.properties.title = "양자 분야 시장·기술·정책·표준화 참고자료 DB"
    wb.properties.creator = "OpenAI"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
